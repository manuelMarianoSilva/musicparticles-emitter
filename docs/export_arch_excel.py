"""
Exports the particle-system architecture diagram to architecture.xlsx.
Produces:
  - Sheet "Graph"  : embedded PNG of the layered architecture diagram
  - Sheet "Nodes"  : table of every node and its layer
  - Sheet "Edges"  : table of every directed edge with its label
"""

import sys
sys.path.insert(0, "/tmp/pylibs")

import io
import textwrap
import networkx as nx
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 1. Graph data ─────────────────────────────────────────────────────────────

LAYERS = {
    "Entry":    ["AndroidManifest", "MainActivity"],
    "UI/GL":    ["ParticleSurfaceView", "ParticleRenderer", "ShaderProgram"],
    "Input":    ["TouchInputManager", "VelocityTracker"],
    "Particle": ["ParticleSystem", "ParticleEmitter", "ParticlePool", "NoiseField", "Particle"],
    "Network":  ["BroadcastServer", "TouchEventSerializer"],
    "Data":     ["TouchEvent", "TouchEventType"],
    "External": ["ExtListeners"],
}

NODE_LABELS = {
    "AndroidManifest":      "AndroidManifest\n(INTERNET, WIFI,\nOpenGL ES 2.0)",
    "MainActivity":         "MainActivity\n(AppCompatActivity)",
    "ParticleSurfaceView":  "ParticleSurfaceView\n(GLSurfaceView)",
    "ParticleRenderer":     "ParticleRenderer\n(GLSurfaceView.Renderer)",
    "ShaderProgram":        "ShaderProgram\n(GLSL vert+frag)",
    "TouchInputManager":    "TouchInputManager",
    "VelocityTracker":      "VelocityTracker\n(Android)",
    "ParticleSystem":       "ParticleSystem\n(cap 20 000)",
    "ParticleEmitter":      "ParticleEmitter",
    "ParticlePool":         "ParticlePool\n(pre-alloc 20 000)",
    "NoiseField":           "NoiseField\n(Perlin noise)",
    "Particle":             "Particle\n(pos, vel, life, hue)",
    "BroadcastServer":      "BroadcastServer\n(UDP :9876 →\n255.255.255.255)",
    "TouchEventSerializer": "TouchEventSerializer\n(JSON ↔ TouchEvent)\n⚠ fromJson() unused",
    "TouchEvent":           "TouchEvent\n(data class)",
    "TouchEventType":       "TouchEventType\n(enum: DOWN/MOVE\n/UP/BURST)",
    "ExtListeners":         "External Listeners\n(LAN :9876)",
}

EDGES = [
    ("AndroidManifest",      "MainActivity",         "launches"),
    ("MainActivity",         "ParticleSurfaceView",  "creates"),
    ("MainActivity",         "BroadcastServer",      "creates"),
    ("MainActivity",         "TouchInputManager",    "creates"),
    ("ParticleSurfaceView",  "ParticleRenderer",     "sets renderer"),
    ("ParticleRenderer",     "ShaderProgram",        "compiles & uses"),
    ("TouchInputManager",    "VelocityTracker",      "uses"),
    ("ParticleSurfaceView",  "TouchInputManager",    "forwards MotionEvent"),
    ("TouchInputManager",    "BroadcastServer",      "emits TouchEvent"),
    ("BroadcastServer",      "TouchEvent",           "enqueues"),
    ("TouchEvent",           "TouchEventType",       "has"),
    ("BroadcastServer",      "TouchEventSerializer", "serializes via"),
    ("BroadcastServer",      "ExtListeners",         "UDP broadcast"),
    ("TouchInputManager",    "ParticleSurfaceView",  "onBurst → queueEvent"),
    ("ParticleSurfaceView",  "ParticleEmitter",      "queueEvent → emit"),
    ("ParticleEmitter",      "ParticleSystem",       "spawn()"),
    ("ParticleSystem",       "ParticlePool",         "manages"),
    ("ParticleSystem",       "NoiseField",           "samples"),
    ("ParticlePool",         "Particle",             "pools"),
    ("NoiseField",           "Particle",             "steers"),
    ("ParticleRenderer",     "ParticleSystem",       "reads liveParticles"),
]

# ── 2. Build networkx graph ───────────────────────────────────────────────────

G = nx.DiGraph()
node_to_layer = {}
for layer, nodes in LAYERS.items():
    for n in nodes:
        G.add_node(n)
        node_to_layer[n] = layer

for src, dst, label in EDGES:
    G.add_edge(src, dst, label=label)

# ── 3. Manual layered layout ─────────────────────────────────────────────────
# Place nodes in columns by layer, evenly spaced vertically.

LAYER_X = {
    "Entry":    0.0,
    "UI/GL":    1.7,
    "Input":    1.7,
    "Particle": 3.4,
    "Network":  3.4,
    "Data":     5.1,
    "External": 5.1,
}
LAYER_Y_START = {
    "Entry":    3.5,
    "UI/GL":    5.0,
    "Input":    2.5,
    "Particle": 5.0,
    "Network":  1.5,
    "Data":     3.5,
    "External": 1.5,
}

pos = {}
for layer, nodes in LAYERS.items():
    x = LAYER_X[layer]
    y0 = LAYER_Y_START[layer]
    for i, n in enumerate(nodes):
        pos[n] = (x, y0 - i * 1.35)

# ── 4. Draw the graph ─────────────────────────────────────────────────────────

LAYER_COLORS = {
    "Entry":    "#4A90D9",
    "UI/GL":    "#7B68EE",
    "Input":    "#50C878",
    "Particle": "#FF8C42",
    "Network":  "#E74C3C",
    "Data":     "#F39C12",
    "External": "#95A5A6",
}

fig, ax = plt.subplots(figsize=(22, 13))
fig.patch.set_facecolor("#0D0D1A")
ax.set_facecolor("#0D0D1A")
ax.axis("off")

# Draw edges first (behind nodes)
for src, dst, data in G.edges(data=True):
    x0, y0 = pos[src]
    x1, y1 = pos[dst]
    ax.annotate(
        "",
        xy=(x1, y1), xytext=(x0, y0),
        arrowprops=dict(
            arrowstyle="-|>",
            color="#888888",
            lw=1.2,
            connectionstyle="arc3,rad=0.08",
        ),
    )
    mx, my = (x0 + x1) / 2, (y0 + y1) / 2
    label = data.get("label", "")
    if label:
        ax.text(mx, my, label, fontsize=5.5, color="#BBBBBB",
                ha="center", va="center",
                bbox=dict(boxstyle="round,pad=0.15", fc="#1A1A2E", ec="none", alpha=0.8))

# Draw nodes
BOX_W, BOX_H = 1.35, 0.90
for node, (x, y) in pos.items():
    layer = node_to_layer[node]
    color = LAYER_COLORS[layer]
    rect = FancyBboxPatch(
        (x - BOX_W / 2, y - BOX_H / 2), BOX_W, BOX_H,
        boxstyle="round,pad=0.05",
        linewidth=1.5, edgecolor=color,
        facecolor=color + "33",   # ~20 % opacity fill
    )
    ax.add_patch(rect)
    wrapped = NODE_LABELS.get(node, node)
    ax.text(x, y, wrapped, fontsize=6.5, color="white",
            ha="center", va="center", fontweight="bold",
            multialignment="center")

# Legend
legend_patches = [
    mpatches.Patch(facecolor=c + "55", edgecolor=c, label=l)
    for l, c in LAYER_COLORS.items()
]
ax.legend(handles=legend_patches, loc="lower right", fontsize=8,
          facecolor="#1A1A2E", edgecolor="#555555", labelcolor="white",
          title="Layers", title_fontsize=8)

ax.set_xlim(-0.9, 6.4)
ax.set_ylim(-1.5, 7.0)
ax.set_title("Particle System — Architecture", color="white", fontsize=14,
             fontweight="bold", pad=12)

plt.tight_layout()

# Save to in-memory PNG
img_buf = io.BytesIO()
fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight",
            facecolor=fig.get_facecolor())
img_buf.seek(0)
plt.close(fig)

# ── 5. Build Excel workbook ───────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── Sheet 1: Graph image ──────────────────────────────────────────────────────
ws_graph = wb.active
ws_graph.title = "Graph"
ws_graph.sheet_view.showGridLines = False
ws_graph["A1"] = "Particle System — Architecture Diagram"
ws_graph["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_graph["A1"].fill = PatternFill("solid", fgColor="0D0D1A")

xl_img = XLImage(img_buf)
xl_img.anchor = "A3"
ws_graph.add_image(xl_img)

# Dark background for graph sheet header row
for col in range(1, 30):
    cell = ws_graph.cell(row=1, column=col)
    cell.fill = PatternFill("solid", fgColor="0D0D1A")

# ── Sheet 2: Nodes ────────────────────────────────────────────────────────────
ws_nodes = wb.create_sheet("Nodes")

header_fill  = PatternFill("solid", fgColor="1E3A5F")
alt_fill     = PatternFill("solid", fgColor="F2F7FC")
header_font  = Font(bold=True, color="FFFFFF", size=11)
border_side  = Side(style="thin", color="CCCCCC")
cell_border  = Border(left=border_side, right=border_side,
                      top=border_side,  bottom=border_side)

headers = ["Node", "Layer", "Description"]
for col, h in enumerate(headers, 1):
    c = ws_nodes.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border

NODE_DESCRIPTIONS = {
    "AndroidManifest":      "Declares permissions (INTERNET, WIFI) and OpenGL ES 2.0 requirement",
    "MainActivity":         "Entry activity; wires GLSurfaceView, TouchInputManager, BroadcastServer",
    "ParticleSurfaceView":  "GLSurfaceView subclass; intercepts touch and drives the emitter",
    "ParticleRenderer":     "GLSurfaceView.Renderer; uploads VBO and calls glDrawArrays each frame",
    "ShaderProgram":        "Compiles GLSL vertex+fragment shaders; provides attrib/uniform handles",
    "TouchInputManager":    "Tracks per-pointer velocity, hold duration, trail length; schedules bursts",
    "VelocityTracker":      "Android system class for computing pointer velocity",
    "ParticleSystem":       "Owns the live particle list; calls update(dt) each frame; thread-safe",
    "ParticleEmitter":      "Translates touch events into particle.spawn() calls with physics params",
    "ParticlePool":         "Pre-allocates 20 000 Particle objects to avoid GC at 60 fps",
    "NoiseField":           "Perlin noise field; returns (nx, ny) steering force per particle per frame",
    "Particle":             "Per-particle state: position, velocity, life, hue, RGB; updated each frame",
    "BroadcastServer":      "Coroutine-based UDP broadcaster; sends every TouchEvent to LAN",
    "TouchEventSerializer": "Converts TouchEvent ↔ JSON. fromJson() is DEAD CODE (never called)",
    "TouchEvent":           "Immutable data class carrying all touch metadata for one pointer event",
    "TouchEventType":       "Enum: TOUCH_DOWN, TOUCH_MOVE, TOUCH_UP, TOUCH_BURST",
    "ExtListeners":         "Any LAN device listening on UDP port 9876",
}

all_nodes = [(n, node_to_layer[n]) for layer_nodes in LAYERS.values() for n in layer_nodes]
for row_idx, (node, layer) in enumerate(all_nodes, 2):
    fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    values = [node, layer, NODE_DESCRIPTIONS.get(node, "")]
    for col, val in enumerate(values, 1):
        c = ws_nodes.cell(row=row_idx, column=col, value=val)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = cell_border

ws_nodes.column_dimensions["A"].width = 26
ws_nodes.column_dimensions["B"].width = 14
ws_nodes.column_dimensions["C"].width = 70
ws_nodes.row_dimensions[1].height = 20

# ── Sheet 3: Edges ────────────────────────────────────────────────────────────
ws_edges = wb.create_sheet("Edges")

headers_e = ["Source", "Target", "Relationship", "Source Layer", "Target Layer"]
for col, h in enumerate(headers_e, 1):
    c = ws_edges.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border

for row_idx, (src, dst, label) in enumerate(EDGES, 2):
    fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    values = [src, dst, label, node_to_layer.get(src, ""), node_to_layer.get(dst, "")]
    for col, val in enumerate(values, 1):
        c = ws_edges.cell(row=row_idx, column=col, value=val)
        c.fill = fill
        c.alignment = Alignment(vertical="top")
        c.border = cell_border

for col, width in zip("ABCDE", [26, 26, 28, 14, 14]):
    ws_edges.column_dimensions[get_column_letter(col.encode()[0] - 64)].width = width
ws_edges.row_dimensions[1].height = 20

# ── 6. Save ───────────────────────────────────────────────────────────────────
out = "/Users/manuelmarianosilva/Documents/particle-system/docs/architecture.xlsx"
wb.save(out)
print(f"Saved: {out}")
